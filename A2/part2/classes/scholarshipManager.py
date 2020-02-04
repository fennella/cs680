from openpyxl import load_workbook
from .scholarship import Scholarship

class ScholarshipManager():

    def __init__(self):

        self._scholarshipList = []
        self._scholarshipFile = load_workbook('dataFiles/dataFile.xlsx')
    
    # Display all scholarships to user
    def displayScholarshipList(self):

        self._loadScholarshipsFromFile()

        for aScholarship in self._scholarshipList:
            print(aScholarship.toString())

    # Write a new scholarship to the Excel sheet
    def addScholarship(self):
        
        desc = input("\nEnter Scholarship Description: \n")
        amount = input("\nEnter Scholarship Amount: \n")

        if not self._validScholarship(desc, amount):
            print("\nInvalid Scholarship Input. Please try again")
            return
        
        scholarship = Scholarship()
        scholarship.setScholarship(desc, amount, self._genScholarshipID())
        self._storeScholarshipToFile(scholarship)

        print(f'\nSuccessfully added scholarship: {scholarship.toString()}\n\n')

    # Load all scholarships from Excel sheet into scholarshipList list
    def _loadScholarshipsFromFile(self):

        for i,row in enumerate(self._scholarshipFile['Scholarships'].iter_rows()):

            if i == 0: continue
        
            scholarship = self._loadScholarshipFromFile(row)
            self._scholarshipList.append(scholarship)
    
    # Initialize the scholarship object
    def _loadScholarshipFromFile(self, row):

        scholarship = Scholarship()
        scholarship.getScholarship(row)
        return scholarship
    
    # When a new scholarship is created, write it to the Excel sheet
    def _storeScholarshipToFile(self, scholarship):

        self._refreshWorkbook()

        newRow = self._scholarshipFile['Scholarships'].max_row + 1

        self._scholarshipFile['Scholarships'].cell(row=newRow, column=1).value = scholarship._id
        self._scholarshipFile['Scholarships'].cell(row=newRow, column=2).value = scholarship._description
        self._scholarshipFile['Scholarships'].cell(row=newRow, column=3).value = scholarship._amount

        self._scholarshipFile.save('dataFiles/dataFile.xlsx')
    
    # Helper function to validate user input
    def _validScholarship(self, desc, amount):

        return len(desc) > 0 and amount.isdigit() and int(amount) > 0

    # Generate a primary key for a scholarship
    def _genScholarshipID(self):

        maxID = 0

        for i,row in enumerate(self._scholarshipFile['Scholarships'].iter_rows()):
            if i == 0: continue

            maxID = max(maxID, int(row[0].value))
        
        return maxID + 1

    # Helper function to refresh workbook when a new student is written to it
    def _refreshWorkbook(self):

        self._scholarshipFile = load_workbook('dataFiles/dataFile.xlsx')



    



    
    
    


