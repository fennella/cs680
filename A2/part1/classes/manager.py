from openpyxl import load_workbook
from .student import Student


class Manager():

    def __init__(self):
        
        self._studentList = []
        self._studentFile = load_workbook('dataFiles/dataFile.xlsx')
    
    # Present all students in the system to the user
    # If studentList has already be created, no need to re-fetch the students
    # When a new student is added it appends it to studentList 
    def displayStudentList(self):

        if len(self._studentList) == 0:
            self._loadStudentsFromFile()

        for aStudent in self._studentList:
            aStudent.displayStudent()

    # Get student input from user then add to studentList and write to Excel sheet
    def addStudent(self):

        first = input("\nEnter Student First Name: ")
        last = input("\nEnter Student Last Name: ")

        if not self._validNames(first, last):
            print("Invalid name entry. Please try again")
            return

        student = Student()
        student.setStudent(first, last)
        self._studentList.append(student)
        self._storeStudentToFile(student)

        print("\nSuccessfully added student: ", end="")
        student.toString()
        print("\n\n")

    # Iterate through all students in Excel sheet and add a student object to studentList for each one
    def _loadStudentsFromFile(self):
        
        for i,row in enumerate(self._studentFile['Students'].iter_rows()):

            if i == 0: continue

            student = self._loadStudentFromFile(row)
            self._studentList.append(student)

    # Create student object from a row in the Excel sheet
    def _loadStudentFromFile(self, row):

        student = Student()
        student.getStudent(row)
        return student

    # Write a new student to the Excel sheet
    # New to refresh the workbook to update it in case user chooses to display students
    # after adding a new one
    def _storeStudentToFile(self, student):

        self._refreshWorkbook()

        newRow = self._studentFile['Students'].max_row + 1

        self._studentFile['Students'].cell(row=newRow, column=1).value = student._studentID
        self._studentFile['Students'].cell(row=newRow, column=2).value = student._first
        self._studentFile['Students'].cell(row=newRow, column=3).value = student._last
        self._studentFile['Students'].cell(row=newRow, column=4).value = student._email

        self._studentFile.save('dataFiles/dataFile.xlsx')


    # This function exists from the UML Model given
    # Do not need this function to meet the functionality requirements
    def _storeStudentsToFile(self):

        print("Store students to file")    

    # Helper function to refresh workbook when a new student is written to it
    def _refreshWorkbook(self):

        self._studentFile = load_workbook('dataFiles/dataFile.xlsx')
    
    # Helper function to validate user input of first and last name 
    def _validNames(self, first, last):

        return len(first) > 0 and len(last) > 0 and first.isalpha() and last.isalpha()

    