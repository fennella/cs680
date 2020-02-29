from openpyxl import load_workbook
from .scholarship import Scholarship
from .dataFile import DataFile

class ScholarshipAdder():

    def __init__(self):

        self._dataFile = DataFile()

    # Write a new scholarship to the Excel sheet
    def add(self):
        
        desc = input("\nEnter Scholarship Description: \n")
        amount = input("\nEnter Scholarship Amount: \n")

        if not self._validScholarship(desc, amount):
            print("\nInvalid Scholarship Input. Please try again")
            return
        
        scholarship = Scholarship()
        scholarship.setScholarship(desc, amount, self._genScholarshipID())
        self._storeScholarship(scholarship)

        print(f'\nSuccessfully added scholarship: {scholarship.toString()}\n\n')


    # When a new scholarship is created, write it to the Excel sheet
    def _storeScholarship(self, scholarship):

        self._dataFile.refreshWorkbook()

        newRow = self._dataFile.file['Scholarships'].max_row + 1

        self._dataFile.file['Scholarships'].cell(row=newRow, column=1).value = scholarship._id
        self._dataFile.file['Scholarships'].cell(row=newRow, column=2).value = scholarship._description
        self._dataFile.file['Scholarships'].cell(row=newRow, column=3).value = scholarship._amount

        self._dataFile.save()

    # Helper function to validate user input
    def _validScholarship(self, desc, amount):

        return len(desc) > 0 and amount.isdigit() and int(amount) > 0

    # Generate a primary key for a scholarship
    def _genScholarshipID(self):

        maxID = 0

        for i,row in enumerate(self._dataFile.file['Scholarships'].iter_rows()):
            if i == 0: continue

            maxID = max(maxID, int(row[0].value))
        
        return maxID + 1

    