from openpyxl import load_workbook

class DataFile():

    def __init__(self):

        self.file = load_workbook('dataFiles/dataFile.xlsx')
    
    def refreshWorkbook(self):

        self.file = load_workbook('dataFiles/dataFile.xlsx')

    def save(self):

        self.file.save('dataFiles/dataFile.xlsx')
    