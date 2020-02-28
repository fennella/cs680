from openpyxl import load_workbook
from .student import Student

class StudentDisplayer():

    def __init__(self):

        self._dataFile = load_workbook('dataFiles/dataFile.xlsx')
        self._studentList = []

    # Present all students in the system to the user
    def display(self):

        self._refreshWorkbook()
        self._loadStudents()

        for aStudent in self._studentList:
            print(aStudent.toString())

    # Iterate through all students in Excel sheet and add a student object to studentList for each one
    def _loadStudents(self):

        self._studentList = []
        
        for i,row in enumerate(self._dataFile['Students'].iter_rows()):

            if i == 0: continue

            student = self._loadStudent(row)
            self._studentList.append(student)

    # Create student object from a row in the Excel sheet
    def _loadStudent(self, row):

        student = Student()
        student.getStudent(row)
        return student 
    
    def _refreshWorkbook(self):

        self._dataFile = load_workbook('dataFiles/dataFile.xlsx')