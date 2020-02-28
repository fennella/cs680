from openpyxl import load_workbook
from .scholarship import Scholarship

class ScholarshipDisplayer():

    def __init__(self):

        self._dataFile = load_workbook('dataFiles/dataFile.xlsx')
        self._scholarshipList = []
    
    # Display all scholarships to user
    def display(self):

        self._refreshWorkbook()
        self._loadScholarships()

        for aScholarship in self._scholarshipList:
            print(aScholarship.toString())

    # Load all scholarships from Excel sheet into scholarshipList list
    def _loadScholarships(self):

        self._scholarshipList = []

        for i,row in enumerate(self._dataFile['Scholarships'].iter_rows()):

            if i == 0: continue
        
            scholarship = self._loadScholarship(row)
            self._scholarshipList.append(scholarship)
    
    # Initialize the scholarship object
    def _loadScholarship(self, row):

        scholarship = Scholarship()
        scholarship.getScholarship(row)
        return scholarship

    def _refreshWorkbook(self):

        self._dataFile = load_workbook('dataFiles/dataFile.xlsx')
