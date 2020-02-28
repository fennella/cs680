from openpyxl import load_workbook
from .student import Student

class StudentAdder():

    def __init__(self):

        self._dataFile = load_workbook('dataFiles/dataFile.xlsx')
    
    # Get student input from user then write to Excel sheet
    def add(self):

        first = input("\nEnter Student First Name: ")
        last = input("\nEnter Student Last Name: ")

        if not self._validNames(first, last):
            print("Invalid name entry. Please try again")
            return

        student = Student()
        student.setStudent(first, last)
        self._storeStudent(student)

        print(f'\nSuccessfully added student: {student.toString()}\n\n')

    # Write a new student to the Excel sheet
    # New to refresh the workbook to update it in case user chooses to display students
    # after adding a new one
    def _storeStudent(self, student):

        self._refreshWorkbook()

        newRow = self._dataFile['Students'].max_row + 1

        self._dataFile['Students'].cell(row=newRow, column=1).value = student._studentID
        self._dataFile['Students'].cell(row=newRow, column=2).value = student._first
        self._dataFile['Students'].cell(row=newRow, column=3).value = student._last
        self._dataFile['Students'].cell(row=newRow, column=4).value = student._email

        self._dataFile.save('dataFiles/dataFile.xlsx')


    # Helper function to validate user input of first and last name 
    def _validNames(self, first, last):

        return len(first) > 0 and len(last) > 0 and first.isalpha() and last.isalpha()

    # Helper function to refresh workbook when a new student is written to it
    def _refreshWorkbook(self):

        self._dataFile = load_workbook('dataFiles/dataFile.xlsx')