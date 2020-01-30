from openpyxl import load_workbook
from .dataFiles import DataFiles

class Scholarship():
    
    # If class is initialized with scholarship ID, automatically know attributes of object
    # If class is not initialized with scholarship ID, ask user for scholarship ID to get attributes
    def __init__(self, scholarshipID=None):

        self._dataFiles = DataFiles()

        if scholarshipID is None:
            self.scholarshipID = self._getScholarshipID()
        else:
            self.scholarshipID = scholarshipID

        attributes = self._getAttributes()
        self.name = attributes[1].value
        self.description = attributes[2].value
        self.studentType = attributes[3].value
        self.dollarAmount = attributes[4].value
        self.type = attributes[5].value
        self.citizenship = attributes[6].value
        self.eligColleges = attributes[7].value
        self.minGPA = float(attributes[8].value)
        self.amountAvailable = int(attributes[9].value)
        self.maxFamilyIncome = int(attributes[10].value)

    # Look in scholarship Excel sheet for scholarship ID supplied
    def _getAttributes(self):
        
        for i,row in enumerate(self._dataFiles.scholarshipSheet.iter_rows()):
            if i == 0: continue
            if int(row[0].value) == self.scholarshipID:
                return row

    # No scholarship ID supplied, ask user for scholarship ID
    def _getScholarshipID(self):

        validScholarshipID = False
        while not validScholarshipID:
            selection = input("Enter the scholarship ID of the scholarship you would like to select: ")
            if not self._isValidScholarshipSelection(selection):
                print("Invalid Selection. Please try again\n")
            else:
                return int(selection)
    
    # Validate scholarship ID is valid
    def _isValidScholarshipSelection(self, selection):

        if not selection.isdigit() or int(selection) < 1 or int(selection) > self._dataFiles.scholarshipSheet.max_row - 1:
            return False
        return True




