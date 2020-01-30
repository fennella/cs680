from openpyxl import load_workbook
from .dataFiles import DataFiles

class Student():

    # If class is initialized with student ID, attributes are known
    # If class is not initialized with student ID, ask user for student ID to get attributes
    def __init__(self, studentID=None):

        self._dataFiles = DataFiles()

        if studentID is None:
            self.studentID = self._getStudentID()
        else: self.studentID = studentID

        attributes = self._getStudentAttributes()
    
        self._firstName = attributes[1].value
        self._lastName = attributes[2].value
        self._email = attributes[3].value
        self.status = attributes[4].value
        self.college = attributes[5].value
        self._major = attributes[6].value
        self.citizenship = attributes[7].value
        self.gpa = attributes[8].value
        self.familyIncome = attributes[9].value

    # Return attributes of student when student ID is known
    def _getStudentAttributes(self):

        for row in self._dataFiles.studentSheet.iter_rows():
            if row[0].value == self.studentID:
                return row
        return False

    # Ask user for student ID to initialize student object
    def _getStudentID(self):

        while True:
            studentID = input("Enter the student ID of the student you would like to select: ")
            if not self._validateStudentID(studentID):
                print("Invalid student ID. Please enter a valid student ID. Please try again.\n")
            else: return studentID
    
    # Check if student ID is in students Excel sheet
    def _validateStudentID(self, studentID):

        if len(studentID) < 1 or len(studentID) > 7: return False
        for row in self._dataFiles.studentSheet.iter_rows():
            if row[0].value == studentID: return True
        return False




