from openpyxl import load_workbook
from .scholarship import Scholarship
from .application import Application
from .award import Award
from .dataFiles import DataFiles

class Admin():

    def __init__(self):

        self._dataFiles = DataFiles()
        self._adminID = self._getAdminID()

        
    def accessScholarshipInfo(self):

        headers = self._dataFiles.scholarshipSheet[1]
        
        print("\n")
        for i,row in enumerate(self._dataFiles.scholarshipSheet.iter_rows()):
            if i == 0: continue
            for j in range(0, len(row)):
                print(f'{headers[j].value}: {row[j].value}')
            print("\n")

    def viewAvailableScholarships(self):

        foundAvailable = False
        print("\n")
        for i, row in enumerate(self._dataFiles.scholarshipSheet.iter_rows()):
            if i == 0: continue
            if int(row[9].value) > 0:
                foundAvailable = True
                print(f'Scholarship ID: {row[0].value}, Name: {row[1].value}, Remaining Scholarships: {row[9].value}')
        if not foundAvailable:
            print("There are no available scholarships to be awared to students")

    def viewEligibleStudents(self):

        self.accessScholarshipInfo()
        
        scholarship = Scholarship()
        if scholarship.type == "Academic":
            validStudents = self._academicFilter(scholarship)
        else: 
            validStudents = self._financialFilter(scholarship)
        
        if len(validStudents) > 0:
            print(f'Eligible Students for the {scholarship.name}\n')
            print("-----------------------------------------------------------")
            for student in validStudents:
                print("\n")
                for attr in student.keys():
                    print(f'{attr}:{student[attr]}', end=", ")
        else:
            print("No students are eligible for this scholarship")

        print("")

    def viewPendingApps(self):
        
        self._dataFiles = DataFiles()

        pendingApps = self._getPendingApps()

        if len(pendingApps) == 0:
            print("No pending student applications found")
            return

        print("")
        for app in pendingApps:
            printStr = ""
            for attr in app.keys():
                printStr += f'{attr}: {app[attr]}, '
            print(printStr[:-2])

        print("")

    def createApplication(self):

        application = Application(self._adminID)
        application.create()

    def declineApplication(self):

    
        rowNum = self._validateAppID()

        self._dataFiles.applicationSheet.cell(row=rowNum, column=4).value = self._adminID
        self._dataFiles.applicationSheet.cell(row=rowNum, column=5).value = "Declined"
        
        self._dataFiles.save()

        print("\nScholarship Successfully Declined\n")

    def awardScholarship(self):

        award = Award(self._adminID)
        award.awardScholarship()

    def removeScholarship(self):

        appID = self._validateAwardID()
        award = Award(self._adminID, appID)
        award.removeScholarship()

    def viewActiveScholarships(self):

        self._dataFiles = DataFiles()

        activeScholarships = self._getActiveScholarships()

        if len(activeScholarships) == 0: 
            print("\nThere are no active scholarships\n")
            return

        
        
        print("")
        for award in activeScholarships:
            printStr = ""
            for attr in award.keys():
                printStr += f'{attr}: {award[attr]}, '
            print(printStr[:-2])

    def _academicFilter(self, scholarship):

        validStudents = []
        for i, row in enumerate(self._dataFiles.studentSheet.iter_rows()):
            if i == 0: continue
            if scholarship.studentType != "All":
                if scholarship.studentType != row[4].value: continue
            if scholarship.citizenship != row[7].value: continue
            if scholarship.minGPA > float(row[8].value): continue
            if scholarship.eligColleges != "All":
                if scholarship.eligColleges != row[5].value: continue
            student = {
                "Student ID":row[0].value,
                "Name":row[1].value + " " + row[2].value,
                "Email":row[3].value,
                "Status":row[4].value,
                "GPA":float(row[8].value)
            }
            validStudents.append(student)

        return sorted(validStudents, key=lambda k: k['GPA'], reverse=True)
                
    def _financialFilter(self, scholarship):
        
        validStudents = []
        for i, row in enumerate(self._dataFiles.studentSheet.iter_rows()):
            if i == 0: continue
            if scholarship.studentType != "All":
                if scholarship.studentType != row[4].value: continue
            if scholarship.citizenship != row[7].value: continue
            if scholarship.eligColleges != "All":
                if scholarship.eligColleges != row[5].value: continue
            if scholarship.maxFamilyIncome < row[9].value: continue
            student = {
                "Student ID":row[0].value,
                "Name":row[1].value + " " + row[2].value,
                "Email":row[3].value,
                "Status":row[4].value,
                "GPA":float(row[8].value)
            }
            validStudents.append(student)

        return sorted(validStudents, key=lambda k: k['GPA'], reverse=True)
        
    def _getAdminID(self):

        while True:

            adminID = input("Enter Your Admin ID: ")
            if len(adminID) < 1 or len(adminID) > 10:
                print("Invalid Admin ID. Please try again\n")
            else:
                return adminID        

    def _getPendingApps(self):

        pendingApps = []
        
        for i,row in enumerate(self._dataFiles.applicationSheet.iter_rows()):
            if i == 0: continue
            if row[4].value == "Pending":
                student = {
                    "Application ID":row[0].value,
                    "Student ID":row[1].value,
                    "Scholarship ID":row[2].value,
                    "Admin ID":row[3].value
                }
                pendingApps.append(student)
        
        return pendingApps

    def _validateAppID(self):

        while True:
            appID = input("Enter Application ID: ")
            if appID.isdigit():
                for i,row in enumerate(self._dataFiles.applicationSheet.iter_rows()):
                    if i == 0: continue
                    if int(row[0].value) == int(appID) and row[4].value == "Pending":
                        return i + 1
            print("No applications with that application ID are pending approval/denial. Please try again\n")

    def _getActiveScholarships(self):

        activeScholarships = []

        for i,row in enumerate(self._dataFiles.awardSheet.iter_rows()):
            if i == 0: continue
            if row[5].value:
                award = {
                    "Award ID":row[0].value,
                    "Student ID":row[2].value,
                    "Scholarship ID":row[1].value,
                    "Date Awarded":row[4].value,
                    "Approved By":row[6].value
                }
                activeScholarships.append(award)

        return activeScholarships

    def _validateAwardID(self):

        while True:
            awardID = input("Enter the award ID of the award you are trying to access: ")
            if awardID.isdigit():
                for i,row in enumerate(self._dataFiles.awardSheet.iter_rows()):
                    if i == 0: continue
                    if int(row[0].value) == int(awardID) and row[5].value:
                        return int(row[3].value)
            
            print("\nInvalid award ID. Please try again\n")

            




        

            

    

