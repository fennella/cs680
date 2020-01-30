from .student import Student
from .scholarship import Scholarship
from .dataFiles import DataFiles
from openpyxl import load_workbook
from datetime import date


class Application():

    # If class is initialized with no application ID,
    # Ask for the scholarship ID and the student ID to create an application ID
    # If the class is initialize with application ID, scholarship and student objects are
    # initialized using this information

    def __init__(self, adminID, applicationID=None):

        self._dataFiles = DataFiles()
        self.adminID = adminID
        self.applicationID = applicationID
        if self.applicationID is None:
            self.scholarship = Scholarship()
            self.student = Student()
        else:
            self.scholarship, self.student = self._getAppInfo(applicationID)
            
    # Check if the user qualifies for a scholarship and then approve/decline application request  
    def create(self):

        if not self._studentQualifies():

            print(f'Student {self.student.studentID} does not qualify for the {self.scholarship.name}\n')
            return

        primaryKey = self._genApplicationPK()
        lastRow = self._dataFiles.applicationSheet.max_row + 1
        
        self._dataFiles.applicationSheet.cell(row=lastRow, column=1).value = primaryKey
        self._dataFiles.applicationSheet.cell(row=lastRow, column=2).value = self.student.studentID
        self._dataFiles.applicationSheet.cell(row=lastRow, column=3).value = self.scholarship.scholarshipID
        self._dataFiles.applicationSheet.cell(row=lastRow, column=4).value = self.adminID
        self._dataFiles.applicationSheet.cell(row=lastRow, column=5).value = "Pending"

        self._dataFiles.save()

        print("\nSuccessfully added application\n")

    # When app ID is supplied by the user, this function initializes the corresponding
    # Scholarship and Student associated with that application
    def _getAppInfo(self, appID):

        for i,row in enumerate(self._dataFiles.applicationSheet.iter_rows()):
            if i == 0: continue
            if int(row[0].value) == int(appID):
                return [Scholarship(int(row[2].value)), Student(row[1].value)]
    
    # Determine if a student qualifies for a scholarship
    def _studentQualifies(self):

        if self.scholarship.amountAvailable < 1: return False

        if self.scholarship.type == "Academic":
            return self._validateAcademicApp(self.student, self.scholarship)
        
        else: return self._validateFinancialApp(self.student, self.scholarship)
    
    # Helper function to verify if a student qualifies for an academic scholarship
    def _validateAcademicApp(self, student, scholarship):

        if scholarship.studentType != "All":
            if scholarship.studentType != student.status: return False
        if scholarship.citizenship != student.citizenship: return False
        if scholarship.eligColleges != "All":
            if scholarship.eligColleges != student.college: return False
        if float(scholarship.minGPA) > float(student.gpa): return False
        
        return True

    # Helper function to verify if a student qualifies for a financial scholarship
    def _validateFinancialApp(self, student, scholarship):

        if scholarship.studentType != "All":
            if scholarship.studentType != student.status: return False
        if scholarship.citizenship != student.citizenship: return False
        if scholarship.eligColleges != "All":
            if scholarship.eligColleges != student.college: return False
        if int(scholarship.maxFamilyIncome) < int(student.familyIncome): return False
        
        return True

    # Helper function to generate a primary key for the application sheet
    def _genApplicationPK(self):

        maxValue = 1
        for i,row in enumerate(self._dataFiles.applicationSheet.iter_rows()):
            if i == 0: continue
            if maxValue < int(row[0].value):
                maxValue = int(row[0].value)

        return maxValue + 1

    