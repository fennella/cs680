from openpyxl import load_workbook
from datetime import date
from .dataFiles import DataFiles
from .application import Application

class Award():

    # If application ID is supplied when initialized, initialze the corresponding application as well
    # If application ID is not supplied, ask user for the corresponding application ID
    def __init__(self, adminID, applicationID=None):
        self._dataFiles = DataFiles()
        self._adminID = adminID
        if applicationID is None:
            self.application = self._getApplication()
        else: self.application = Application(self._adminID, applicationID)

    # If scholarship has availability, award to a student
    def awardScholarship(self):

        if self.application.scholarship.amountAvailable < 1:
            print("Unable to award scholarship. All scholarships are currently taken")
            return 
        
        # Update award Excel sheet
        self._writeNewAward()
        # Update application Excel sheet
        self._updateApplication()
        # Decrement amount available in scholarship Excel sheet
        self._decrementScholarship()

        print("\nScholarship successfully awarded\n")
    
    # Remove scholarship from student
    def removeScholarship(self):

        rowNum = self._getAwardRow()
        self._dataFiles.awardSheet.cell(row=rowNum, column=6).value = False
        self._dataFiles.save()
        # Increment amount available in scholarship Excel sheet
        self._incrementScholarship()

        print("\nScholarship successfully removed\n")

    # Helper function to create a new entry in the award Excel sheet
    def _writeNewAward(self):

        awardID = self._genAwardPK()
        lastRow = self._dataFiles.awardSheet.max_row + 1

        self._dataFiles.awardSheet.cell(row=lastRow, column=1).value = awardID
        self._dataFiles.awardSheet.cell(row=lastRow, column=2).value = self.application.scholarship.scholarshipID
        self._dataFiles.awardSheet.cell(row=lastRow, column=3).value = self.application.student.studentID
        self._dataFiles.awardSheet.cell(row=lastRow, column=4).value = self.application.applicationID
        self._dataFiles.awardSheet.cell(row=lastRow, column=5).value = date.today()
        self._dataFiles.awardSheet.cell(row=lastRow, column=6).value = True
        self._dataFiles.awardSheet.cell(row=lastRow, column=7).value = self._adminID

        self._dataFiles.save()

    # Find the award associated with the given application ID
    def _getAwardRow(self):

        for i,row in enumerate(self._dataFiles.awardSheet.iter_rows()):
            if i == 0: continue
            if int(row[3].value) == int(self.application.applicationID):
                return i + 1

    # When scholarship gets awarded, decrement amount available in the scholarship Excel sheet
    def _decrementScholarship(self):

        for i,row in enumerate(self._dataFiles.scholarshipSheet.iter_rows()):
            if i == 0: continue
            if int(row[0].value) == int(self.application.scholarship.scholarshipID):
                currAmount = int(self._dataFiles.scholarshipSheet.cell(row=i + 1, column=10).value)
                self._dataFiles.scholarshipSheet.cell(row=i + 1, column=10).value = currAmount - 1
                self._dataFiles.save()
                return

    # When scholarship gets removed, increment amount available in scholarship Excel sheet
    def _incrementScholarship(self):

        for i,row in enumerate(self._dataFiles.scholarshipSheet.iter_rows()):
            if i == 0: continue
            if int(row[0].value) == int(self.application.scholarship.scholarshipID):
                currAmount = int(self._dataFiles.scholarshipSheet.cell(row=i + 1, column=10).value)
                self._dataFiles.scholarshipSheet.cell(row=i + 1, column=10).value = currAmount + 1
                self._dataFiles.save()
                return

    # When scholarship gets approved, update status in application Excel sheet
    def _updateApplication(self):

        for i,row in enumerate(self._dataFiles.applicationSheet.iter_rows()):
            if i == 0: continue
            if int(row[0].value) == int(self.application.applicationID):
                self._dataFiles.applicationSheet.cell(row=i + 1, column=5).value = "Approved"
                self._dataFiles.save()
                return

    # Ask user for application ID and then validate if its a valid application ID
    def _getApplication(self):

        while True:
            appID = input("Enter the application ID for the application you would like to access: ")
            if self._validateAppID(appID): return Application(self._adminID, appID)
            else: print("Invalid Application ID. Please try again")
    
    # Check if user supplied application ID is a valid application ID
    def _validateAppID(self, appID):

        if not appID.isdigit(): return False

        for i,row in enumerate(self._dataFiles.applicationSheet.iter_rows()):
            if i == 0: continue
            if int(row[0].value) == int(appID) and row[4].value == "Pending":
                return True
        return False

    # Generate an award primary key to be used in award Excel sheet
    def _genAwardPK(self):

        maxValue = 1
        for i,row in enumerate(self._dataFiles.awardSheet.iter_rows()):
            if i == 0: continue
            if maxValue < int(row[0].value):
                maxValue = int(row[0].value)

        return maxValue + 1





