from openpyxl import load_workbook
import names
import random

STUDENT_STATUSES = ["Undergrad", "Grad", "Both"]

def createHeaders(sheet):

    
    headers = ["StudentID", "FirstName", "LastName", "Email", "Status", "College", "Major", "Citizenship", "GPA", "FamilyIncome"]
    for i,header in enumerate(headers):
        sheet.cell(row=1, column=i + 1).value = header

def loadCollegeData():

    collegeWB = load_workbook('collegeMajors.xlsx')
    return collegeWB['College']



def createStudentFile(wb):

    sheet = wb['Students']

    createHeaders(sheet)
    collegeData = loadCollegeData()

    for row in range(2, 5000):

        firstName = names.get_first_name()
        lastName = names.get_last_name()
        studentID = firstName[0].lower() + lastName[0].lower() + str(random.randint(1, 999))
        email = studentID + "@drexel.edu"
        status = STUDENT_STATUSES[random.randint(0,2)]
        randRow = random.randint(1, 174)
        major = collegeData["A" + str(randRow)].value
        college = collegeData["B" + str(randRow)].value
        if randRow % 2 == 0 or randRow < 44: citizenship = True
        else: citizenship = False
        gpa = round(random.uniform(2.5, 4.0), 2)
        familyIncome = random.randint(15000, 500000)
        if familyIncome % 2 == 0: familyIncome = familyIncome / 2

        data = [studentID, firstName, lastName, email, status, college, major, citizenship, gpa, familyIncome]

        for i,value in enumerate(data):
            sheet.cell(row=row, column=i + 1).value = value


    wb.save('dataFile.xlsx')


def main():

    wb = load_workbook('dataFile.xlsx')
    createStudentFile(wb)


main()




